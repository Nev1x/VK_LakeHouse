"""XLSX-ридер: openpyxl read_only+data_only iter_rows (НЕ pandas.read_excel — грузит книгу целиком).

Каждый непустой лист = отдельный источник `<источник>_<лист>`. Merged cells: значение в
левой-верхней ячейке, остальные None (openpyxl в read_only отдаёт None для не-anchor ячеек).
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from loftnav.config import IngestConfig
from loftnav.ingest.readers.base import Source, normalize_cell


class ExcelReader:
    @classmethod
    def can_read(cls, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def __init__(self, cfg: IngestConfig) -> None:
        self._cfg = cfg

    def sources(self, path: Path) -> Iterator[Source]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                header = _first_nonempty(rows)
                if header is None:
                    continue  # пустой лист пропускаем
                # сырые имена (str|None); нормализацию col_N/_2 делает sanitize_columns (INFO-1)
                raw_header = [normalize_cell(c) for c in header]
                yield Source(
                    suffix=ws.title,
                    records=_sheet_records(len(raw_header), rows),
                    header=raw_header,
                )
        finally:
            wb.close()


def _first_nonempty(rows: Iterator[tuple]) -> tuple | None:
    for row in rows:
        if any(c is not None and str(c).strip() != "" for c in row):
            return row
    return None


def _sheet_records(ncol: int, rows: Iterator[tuple]) -> Iterator[list[str | None]]:
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # пустая строка листа
        yield [normalize_cell(row[i]) if i < len(row) else None for i in range(ncol)]
