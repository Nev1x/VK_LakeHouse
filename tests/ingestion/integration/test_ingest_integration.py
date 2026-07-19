"""Интеграционные тесты ingestion на живом стеке (FR-010/FR-012/FR-015). Маркер requires_stack.

Контент фикстур уникален (uuid): идемпотентность по content-hash глобальна, фиксированный контент
коллизировал бы между прогонами. Каждый тест убирает свои таблицы за собой.
"""

from __future__ import annotations

import dataclasses
import json
import uuid
from pathlib import Path

import pytest
from openpyxl import Workbook

from loftnav.config import IngestConfig
from loftnav.ingest import bronze_writer
from loftnav.ingest.run import EXIT_ALL_FAILED, EXIT_OK, EXIT_PARTIAL, ingest_paths
from loftnav.trino_client import get_connection

pytestmark = pytest.mark.requires_stack


def _q(sql: str):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(sql)
        return cur.fetchall()
    finally:
        conn.close()


def _count(table: str) -> int:
    return _q(f"SELECT count(*) FROM {table}")[0][0]


def _drop(*sources: str) -> None:
    conn = get_connection()
    try:
        cur = conn.cursor()
        for source in sources:
            for sql in (
                f'DROP TABLE IF EXISTS iceberg.bronze."{source}"',
                f'DROP TABLE IF EXISTS iceberg.quarantine."bronze_{source}_rejects"',
            ):
                cur.execute(sql)
                cur.fetchall()
    finally:
        conn.close()


def test_demo_multiformat_and_broken(tmp_path: Path) -> None:
    """≥3 формата/схемы загружены одной командой; битый -> exit 2, журнал failed (крит.1/3/5)."""
    tag = uuid.uuid4().hex[:8]
    d = tmp_path / "demo"
    d.mkdir()

    csv = d / f"apt_{tag}.csv"  # cp1251, ';', запятая-десятичная, _note -> u_note
    csv.write_bytes(
        ("\r\n".join([
            "id;city;area;price;_note",
            f"1;Москва;45,5;5000000;{tag}",
            "2;Казань;38,2;3100000;",
            "3;Сочи;72,1;9800000;элит",
        ]) + "\r\n").encode("cp1251")
    )

    jsonl = d / f"flats_{tag}.jsonl"  # вложенный объект
    with open(jsonl, "w", encoding="utf-8") as f:
        for i in range(2):
            f.write(json.dumps({"id": i, "loc": {"lat": 55.0 + i}, "tag": tag}) + "\n")

    xlsx = d / f"book_{tag}.xlsx"  # 2 листа + merged cell
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "flats"
    ws1.append(["id", "rooms", "price"])
    ws1.append([1, 2, 100])
    ws1.append([2, 3, 200])
    ws1.merge_cells("B4:C4")
    ws1["A4"] = 3
    ws1["B4"] = "merged"
    ws2 = wb.create_sheet("meta")
    ws2.append(["k", "v"])
    ws2.append(["src", tag])
    wb.save(xlsx)

    broken = d / f"broken_{tag}.csv"
    broken.write_bytes(bytes([0, 1, 2, 3, 250, 0] * 300))

    src = [f"apt_{tag}", f"flats_{tag}", f"book_{tag}_flats", f"book_{tag}_meta"]
    try:
        code = ingest_paths([d], None, IngestConfig.from_env())
        assert code == EXIT_PARTIAL
        assert _count(f'iceberg.bronze."{src[0]}"') == 3
        assert _count(f'iceberg.bronze."{src[1]}"') == 2
        assert _count(f'iceberg.bronze."{src[2]}"') == 3  # 2 обычные + 1 merged-row
        assert _count(f'iceberg.bronze."{src[3]}"') == 1
        # u_note присутствует (переименование пользовательского _-префикса)
        cols = [r[0] for r in _q(f'DESCRIBE iceberg.bronze."{src[0]}"')]
        assert "u_note" in cols and "_run_id" in cols
        # журнал: битый -> failed
        st = _q(
            "SELECT source_file, status FROM iceberg.ops.pipeline_runs "
            f"WHERE source_file LIKE '%broken_{tag}.csv'"
        )
        assert st and st[0][1] == "failed"
    finally:
        _drop(*src)


def test_idempotent_repeat_no_growth(tmp_path: Path) -> None:
    """Повторный ingest неизменённого файла -> журнал skipped, счётчик не растёт (крит.2, US-3)."""
    tag = uuid.uuid4().hex[:8]
    source = f"idem_{tag}"
    csv = tmp_path / f"{source}.csv"
    csv.write_text(f"id,name\n1,a-{tag}\n2,b\n3,c\n", encoding="utf-8")
    cfg = IngestConfig.from_env()
    try:
        assert ingest_paths([csv], source, cfg) == EXIT_OK
        assert _count(f'iceberg.bronze."{source}"') == 3
        assert ingest_paths([csv], source, cfg) == EXIT_OK  # повтор
        assert _count(f'iceberg.bronze."{source}"') == 3     # без роста
        last = _q(
            "SELECT status FROM iceberg.ops.pipeline_runs "
            f"WHERE source_file = '{csv}' ORDER BY started_at DESC LIMIT 1"
        )[0][0]
        assert last == "skipped"
    finally:
        _drop(source)


def test_quarantine_and_sum(tmp_path: Path) -> None:
    """Невалидные строки -> quarantine; rows_ok+rows_q = число строк источника (крит.4, US-5)."""
    tag = uuid.uuid4().hex[:8]
    source = f"qtest_{tag}"
    csv = tmp_path / f"{source}.csv"
    # буфер инференса = первые 2 строки -> val BIGINT; 'oops' в 3-й -> reject
    csv.write_text(f"id,val\n1,10\n2,20\n3,oops-{tag}\n4,40\n", encoding="utf-8")
    cfg = dataclasses.replace(IngestConfig.from_env(), read_chunk_rows=2)
    try:
        assert ingest_paths([csv], source, cfg) == EXIT_OK
        bronze = _count(f'iceberg.bronze."{source}"')
        rejects = _count(f'iceberg.quarantine."bronze_{source}_rejects"')
        assert (bronze, rejects) == (3, 1)
        assert bronze + rejects == 4  # = число строк источника
        jr = _q(
            "SELECT rows_ok, rows_quarantined FROM iceberg.ops.pipeline_runs "
            f"WHERE source_file = '{csv}' ORDER BY started_at DESC LIMIT 1"
        )[0]
        assert (jr[0], jr[1]) == (3, 1)
    finally:
        _drop(source)


def test_replay_no_duplication(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Искусственный failed-прогон (сбой после вставки) -> replay -> строки не задвоились."""
    tag = uuid.uuid4().hex[:8]
    source = f"rtest_{tag}"
    csv = tmp_path / f"{source}.csv"
    csv.write_text(f"id,val\n1,a-{tag}\n2,b\n3,c\n", encoding="utf-8")
    cfg = IngestConfig.from_env()
    original = bronze_writer.insert_rows

    def failing(*args, **kwargs):
        original(*args, **kwargs)  # строки реально вставлены (committed) до сбоя
        raise RuntimeError("искусственный сбой после вставки")

    try:
        monkeypatch.setattr(bronze_writer, "insert_rows", failing)
        assert ingest_paths([csv], source, cfg) == EXIT_ALL_FAILED
        assert _count(f'iceberg.bronze."{source}"') == 3  # частичные строки крашнутого прогона

        monkeypatch.setattr(bronze_writer, "insert_rows", original)  # чиним
        assert ingest_paths([csv], source, cfg) == EXIT_OK
        # replay удалил строки failed-прогона и вставил заново -> ровно 3, без задвоения
        assert _count(f'iceberg.bronze."{source}"') == 3
    finally:
        _drop(source)
